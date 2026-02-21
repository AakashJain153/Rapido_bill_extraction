import os
import re
import shutil
import pdfplumber
import pandas as pd
from datetime import datetime
from tkinter import Tk
from tkinter.filedialog import askdirectory
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ---------------------------------------
# EXTRACT DETAILS FUNCTION
# ---------------------------------------

def extract_details(pdf_path):

    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            extracted = page.extract_text()
            if extracted:
                text += extracted + "\n"

    lines = [l.strip() for l in text.split("\n") if l.strip()]

    # ---------------- DATE ----------------
    ride_date = None
    date_match = re.search(
        r'([A-Za-z]+ \d{1,2}(st|nd|rd|th)? \d{4}, \d{1,2}:\d{2} [APM]{2})',
        text
    )

    if date_match:
        date_str = re.sub(r'(st|nd|rd|th)', '', date_match.group(1))
        try:
            ride_date = datetime.strptime(date_str, "%b %d %Y, %I:%M %p")
        except:
            pass

    # ---------------- RIDE ID ----------------
    ride_id_match = re.search(r'RD\d+', text)
    ride_id = ride_id_match.group(0) if ride_id_match else ""



    # ---------------- VEHICLE NUMBER ----------------
    vehicle_number = ""
    plate_pattern = re.compile(r'[A-Z]{2}[0-9]{2}[A-Z]{1,2}[0-9]{4}')
    for line in lines:
        cleaned = re.sub(r'[^A-Za-z0-9]', '', line).upper()
        match = plate_pattern.search(cleaned)
        if match:
            vehicle_number = match.group(0)
            break

    # ---------------- FARE ----------------
    fare_amount = 0.0
    fare_match = re.search(r'₹\s*([\d,]+)', text)
    if fare_match:
        fare_amount = float(fare_match.group(1).replace(",", ""))

    # ---------------- PICKUP & DROP (Robust Multi-line Handling) ----------------
    pickup = ""
    drop = ""

    fare_index = None
    for i, line in enumerate(lines):
        if "₹" in line:
            fare_index = i
            break

    if fare_index is not None:

        addresses = []
        current_address = []

        for i in range(fare_index + 1, len(lines)):

            line = lines[i].strip()

            # Stop at disclaimer
            if line.lower().startswith("this document"):
                break

            # Build address lines
            current_address.append(line)

            # If line ends with India → complete one address
            if line.lower().endswith("india"):
                full_address = " ".join(current_address)
                addresses.append(full_address)
                current_address = []

        if len(addresses) >= 1:
            pickup = addresses[0]

        if len(addresses) >= 2:
            drop = addresses[1]


    return {
        "Original File": os.path.basename(pdf_path),
        "Date": ride_date,
        "Ride ID": ride_id,
        "Vehicle Number": vehicle_number,
        "Pickup Location": pickup,
        "Drop Location": drop,
        "Fare Amount": fare_amount,
        "Full Path": pdf_path
    }


# ---------------------------------------
# MAIN FUNCTION
# ---------------------------------------

def main():

    # Folder popup
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    folder_path = askdirectory(title="Select Folder Containing Rapido PDF Bills")

    if not folder_path:
        print("No folder selected. Exiting.")
        return

    refined_folder = os.path.join(folder_path, "Refined")
    os.makedirs(refined_folder, exist_ok=True)

    records = []

    # Only process PDFs in root folder (not Refined folder)
    pdf_files = []
    for f in os.listdir(folder_path):
        full_path = os.path.join(folder_path, f)
        if os.path.isfile(full_path) and f.lower().endswith(".pdf"):
            pdf_files.append(f)

    if not pdf_files:
        print("⚠ No PDF files found in selected folder.")
        return

    for file in pdf_files:
        full_path = os.path.join(folder_path, file)

        try:
            data = extract_details(full_path)

            # Create renamed COPY (not move)
            if data["Date"] and data["Fare Amount"] > 0:

                new_filename = f"{data['Date'].strftime('%Y%m%d')}_{data['Fare Amount']:.2f}.pdf"
                new_path = os.path.join(refined_folder, new_filename)

                # Avoid overwrite
                counter = 1
                while os.path.exists(new_path):
                    new_filename = f"{data['Date'].strftime('%Y%m%d')}_{data['Fare Amount']:.2f}_{counter}.pdf"
                    new_path = os.path.join(refined_folder, new_filename)
                    counter += 1

                # Copy instead of rename (original remains untouched)
                shutil.copy2(full_path, new_path)

                data["Full Path"] = new_path
                data["Original File"] = new_filename

            records.append(data)

        except Exception as e:
            print(f"Error processing {file}: {e}")

    df = pd.DataFrame(records)

    if df.empty:
        print("⚠ No valid Rapido PDFs were processed.")
        return

    excel_path = os.path.join(refined_folder, "Rapido_Bills_Summary.xlsx")

    df_to_save = df.drop(columns=["Full Path"], errors="ignore")
    df_to_save.to_excel(excel_path, index=False)

    # Add hyperlinks
    wb = load_workbook(excel_path)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        file_path = df.iloc[row - 2]["Full Path"]
        ws.cell(row=row, column=1).hyperlink = file_path
        ws.cell(row=row, column=1).style = "Hyperlink"

    # Auto column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(column)].width = min(max_length + 2, 80)

    wb.save(excel_path)

    print("\n✅ Extraction Complete")
    print("Original PDFs were NOT modified.")
    print(f"Refined copies saved in:\n{refined_folder}")
    print(f"Excel summary saved at:\n{excel_path}")


if __name__ == "__main__":
    main()
